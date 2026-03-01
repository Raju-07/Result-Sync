# Standard Library Imports
import os
import sys
import platform
import subprocess
import threading
import webbrowser
from tkinter import filedialog
from tkinter import messagebox as msg
from tkinter import END

# GUI Imports
from customtkinter import *
from PIL import Image

# Excel Imports
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

# Web Scraping Imports
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException

class AppResult:
    def __init__(self):
        # Setting custom appearance
        set_default_color_theme("dark-blue")
        set_appearance_mode('dark')

        # Creating Window
        self.app = CTk()
        self.app.geometry('1000x600')
        self.app.title("Result Sync")
        self.app.configure(fg_color="#0F172A")    

        # Icon Setup
        self.file_path = os.path.dirname(os.path.realpath(__file__))        
        self.folder_icon = CTkImage(Image.open(self.resource_path("icon.png")), size=(33, 33))

        # Main Interface Frame
        self.frame1 = CTkFrame(master=self.app, width=310, height=450, fg_color="#111827", corner_radius=10, border_color='#1F2937', border_width=2)
        self.frame1.place(relx=0.83, rely=0.55, anchor=CENTER)
        
        # Adding Labels to the application
        self.create_label(self.app, "Result Sync", 0.5, 0.06, font=("Eras Bold ITC", 50, "bold"), text_color="#F1F5F9")
        self.create_label(self.app, "━━━━ Automate Result Extraction & Excel Reporting ━━━━", 0.5, 0.13, text_color="#CBD5E1")
        self.create_label(self.app, "©2025 Raju Yadav | Powered by Python, made with Love ❤", 0.5, 0.97, text_color="#94A3B8")

        self.create_label(self.app, "Status", 0.4, 0.33, ("Eras Bold ITC", 22, "bold"))
        self.current_student = self.create_label(self.app, "Not Started Yet", 0.4, 0.37, ("Roboto", 18, "normal"), text_color="#1ACCF8")

        # Frame 1 Widgets
        self.create_label(self.frame1, "Required Info", 0.5, 0.07, font=("Eras Bold ITC", 27, "bold"))

        self.create_label(self.frame1, "Select Record File (.xlsx)", 0.33, 0.2, font=("Roboto", 16, "normal"))
        self.input_file = self.create_entrybox(self.frame1, "Input Record File Path", 0.42, 0.28, 230)
        self.create_button(self.frame1, "", 0.9, 0.28, self.select_record_file, width=10, height=10, image=self.folder_icon, fg_color="#1F2937", hover_color="#374151")

        self.create_label(self.frame1, "Location to Save", 0.26, 0.4, font=("Roboto", 16, "normal"))
        self.output_file = self.create_entrybox(self.frame1, "Select Where to Save", 0.42, 0.48, 230)
        self.create_button(self.frame1, "", 0.9, 0.48, self.select_loc_name, width=10, height=10, image=self.folder_icon, fg_color="#1F2937", hover_color="#374151")
        
        self.create_button(self.frame1, "Proceed", 0.5, 0.65, self.proceed, font=("Roboto", 18, "bold"))
        
        # Important Notes
        self.create_label(self.frame1, "NOTE: FIRST ROW IS TREATED AS \nHEADERS. REGISTRATION NO. MUST BE \nIN COLUMN A & ROLL NO. IN COLUMN B", 0.5, 0.8, font=("calibri", 12, "normal"), text_color='#F87171')
        self.create_label(self.frame1, "NOTE: IF AN ERROR OCCURS, RESELECT\nTHE SAME OUTPUT FILE", 0.5, 0.93, font=("calibri", 12, "normal"), text_color='#F87171')

        self.create_label(self.app, "Please provide your valuable feedback", 0.13, 0.85, font=("Roboto", 12, "normal"))
        self.feedback_lbl = self.create_label(self.app, "Click Here", 0.1, 0.9, font=("Roboto", 12, "normal"))
        self.feedback_lbl.bind("<Button-1>", lambda event: self.open_feedback())
        self.feedback_lbl.bind("<Enter>", lambda event: self.feedback_lbl.configure(font=("Roboto", 14, "underline"), cursor="hand2"))
        self.feedback_lbl.bind("<Leave>", lambda event: self.feedback_lbl.configure(font=("Roboto", 12, "normal"), cursor="arrow"))

        self.app.mainloop()

    # Making path more reliable
    def resource_path(self, relative_path):
        """Get absolute path to resource, works for dev and for PyInstaller"""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    # Feedback form for improving the project
    def open_feedback(self):
        webbrowser.open('https://forms.gle/PoXZRNdXS1P14MsaA')
    
    # Reuseable Components
    def create_label(self, master, text:str, x:float, y:float, font=("Roboto", 18, "normal"), text_color=None):
        label = CTkLabel(master=master, text=text, font=font, text_color=text_color)
        label.place(relx=x, rely=y, anchor=CENTER)
        return label
    
    def create_button(self, master, text:str, relx:float, rely:float, command, width=150, height=35, font=("Roboto", 16, "normal"), image=None, compound=LEFT, fg_color="#22C55E", hover_color="#16A34A"):
        button = CTkButton(master=master, width=width, height=height, corner_radius=8, font=font, image=image, compound=compound, text=text, command=command, fg_color=fg_color, hover_color=hover_color)
        button.place(relx=relx, rely=rely, anchor=CENTER)
        return button
    
    def create_entrybox(self, master, placeholder_text, relx:float, rely:float, width=180, height=35, cor_rad=6, bdr_clr='#1F2937', bdr_wid=2, font=("Roboto", 14, "normal")):
        entrybox = CTkEntry(master=master, placeholder_text=placeholder_text, corner_radius=cor_rad, width=width, height=height, border_color=bdr_clr, border_width=bdr_wid, font=font, fg_color="#020617", text_color="#E5E7EB")
        entrybox.place(relx=relx, rely=rely, anchor=CENTER)
        return entrybox
    
    # geting total record count
    def get_total_records(self):
        if os.path.exists(self.input_file.get()):
            wb = load_workbook(self.input_file.get())
            ws = wb.active
            row_count = 0
            for rows in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if rows and rows[0] is not None:
                    row_count += 1
            return row_count
        else:
            msg.showerror("File Not Exists", "Please provide the correct path and location.")
            return 0

    def proceed(self):
        if self.input_file.get() and self.output_file.get():
            self.input_file.configure(border_color='green')
            self.output_file.configure(border_color="green")

            self.estimate_time = self.get_total_records()
            if self.estimate_time == 0:
                return

            self.progress_status = 0.0
            self.progressbar = CTkProgressBar(self.app, 250, height=18, corner_radius=6, border_color="blue", progress_color="#34D399", orientation="horizontal")
            self.progressbar.place(relx=0.42, rely=0.5, anchor=CENTER)
            self.progressbar.set(self.progress_status) 
            
            self.percentage_label = self.create_label(self.app, "0.00%", 0.6, 0.5)
            self.est_time_label = self.create_label(self.app, f"Estimated Time: {self.estimate_time} Minutes", 0.43, 0.55)

            self.headless_var = BooleanVar(value=False) 
            self.headless_switch = CTkSwitch(self.app, text="Headless Mode (Faster)  ", variable=self.headless_var, progress_color="#34D399")
            self.headless_switch.place(relx=0.40, rely=0.62, anchor=CENTER)

            self.openfile_var = BooleanVar(value=True)
            self.openfile_switch = CTkSwitch(self.app,text="Open file when completed",variable=self.openfile_var,progress_color="#34D399")
            self.openfile_switch.place(relx=0.40,rely=0.66,anchor= CENTER)
            
            self.start_btn = self.create_button(self.app, "Start", 0.41, 0.75, command=lambda: threading.Thread(target=self.run_automation,args=(True,), daemon=True).start())

        else:
            self.input_file.configure(border_color="green" if self.input_file.get() else "red")
            self.output_file.configure(border_color="green" if self.output_file.get() else "red")
            msg.showerror("File Not Selected", "Please select both input and output files.")

    def select_record_file(self):
        file_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Select Record File")
        if file_path:
            self.input_file.configure(state="normal")
            self.input_file.delete(0, END)
            self.input_file.insert(0, file_path)
            self.input_file.configure(state="readonly")

    def select_loc_name(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Select Where to Save Output File")
        if file_path:
            self.output_file.configure(state="normal")
            self.output_file.delete(0, END)        
            self.output_file.insert(0, file_path)
            self.output_file.configure(state="readonly")

    def get_row_tracker(self):
        self.current_row = 4
        self.row_tracker_file = "row_status.txt"
        if os.path.exists(self.row_tracker_file):
            with open(self.row_tracker_file, 'r') as file:
                row = file.readline().strip()
                if row.isdigit():
                    return int(row)
        
        with open(self.row_tracker_file, 'w') as file:
            file.write(str(self.current_row))
        return self.current_row

    def run_automation(self,format=False):
        self.format = format
        if self.input_file.get():
            input_wb = load_workbook(self.input_file.get())
            input_ws = input_wb.active
        else:
            msg.showerror("File Not Loaded", "Error during loading input file.")
            return

        self.current_row = self.get_row_tracker()
        total = self.get_total_records()
        
        if total == 0:
            msg.showerror("No Data", "Excel file has no records.")
            return
            
        self.one_time_increment = 1 / total

        # Initialize Output Excel File
        if os.path.exists(self.output_file.get()):
            wb = load_workbook(self.output_file.get())
        else:
            wb = Workbook()

        ws = wb['Sheet1'] if "Sheet1" in wb.sheetnames else wb.active
        ws.title = "Sheet1"

        # Headers Formatting
        ws.merge_cells('A1:N1')
        ws['A1'].fill = PatternFill(start_color="8DFAB1", end_color="8DFAB1", fill_type="solid")
        ws['A1'] = f"BCA Result"
        ws['A1'].font = Font("Times New Roman", size=16, bold=True, color="006100")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

        section_headers = [
            {"cell": "A2:E2", "text": "Students Details"},
            {"cell": "G2:K2", "text": "Marks in Each Subject"},
            {"cell": "M2:N2", "text": "Result"}
        ]

        for section in section_headers:
            ws.merge_cells(section["cell"])
            cell_ref = section["cell"].split(":")[0]
            ws[cell_ref] = section["text"]
            ws[cell_ref].fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            ws[cell_ref].font = Font(name="Century", size=14, color="006100")
            ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")

        col_head = ["Sr.No", "Name", "Father Name", "Registration No", "Roll No", "", "Sub_1", "Sub_2", "Sub_3", "Sub_4", "Sub_5", "", "Total Marks", "Result"]
        for cols, header in enumerate(col_head, start=1):
            cols_char = get_column_letter(cols)
            ws[f"{cols_char}3"] = header

        # Webdriver Setup
        options = Options()
        if self.headless_var.get():
            options.add_argument("--headless=new")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-gpu")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")        
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        try:
            driver = webdriver.Chrome(options=options)
        except Exception:
            msg.showerror("Browser Error", "Chrome not installed or blocked. Install Google Chrome.")
            return

        wait = WebDriverWait(driver, 10)
        
        try:
            driver.get("https://result.mdu.ac.in/postexam/result.aspx")
        except Exception:
            msg.showerror("Internet Error", "Check your internet connection.")
            driver.quit()
            return

        self.process_data = []
        file_name = "Processed_entries.txt"
        if os.path.exists(file_name):
            try:
                with open(file_name, "r") as file:
                    self.process_data = list(map(str.strip, file.readlines()))
            except Exception:
                msg.showerror("Error", "Error reading Processed Entries.")

        try:
            with open(file_name, "a") as file:
                for rows in input_ws.iter_rows(min_row=2, max_col=2, values_only=True):
                    if rows[0] is None:
                        break
                    
                    reg_no = str(rows[0]).strip()
                    if reg_no in self.process_data:
                        continue

                    # Extract Result
                    reg_input = wait.until(EC.presence_of_element_located((By.ID, "txtRegistrationNo")))
                    reg_input.send_keys(str(reg_no))
                    driver.find_element(By.ID, "txtRollNo").send_keys(str(rows[1]).strip())
                    wait.until(EC.element_to_be_clickable((By.ID, "cmdbtnProceed"))).click()
                    wait.until(EC.element_to_be_clickable((By.ID, "imgComfirm"))).click()
                    wait.until(EC.element_to_be_clickable((By.ID, "rptMain_ctl01_lnkView"))).click()
                    wait.until(EC.visibility_of_element_located((By.ID, "lblStudentName")))
                    
                    student_name = driver.find_element(By.ID, "lblStudentName").text
                    father_name = driver.find_element(By.ID, "lblFatherName").text
                    self.current_student.configure(text=student_name)
                    #updating the Excel with Semter and Subject
                    if self.format:
                        sem_header = driver.find_element(By.ID,"lblSem").text
                        sub_name = [driver.find_element(By.XPATH, f"//table[contains(@class, 'table_mdm')]//tr[{i}]//td[1]").text.strip(" :") for i in range(2, 7)]
                        sub_name1,sub_name2,sub_name3,sub_name4,sub_name5 = sub_name
                        # print(f"{sub_name1 = } {sub_name2 = } {sub_name3 = } {sub_name4 = } {sub_name5 = }")

                    subject_marks = [driver.find_element(By.XPATH, f"//table[contains(@class, 'table_mdm')]//tr[{i}]//td[8]").text for i in range(2, 7)]
                    sub_1, sub_2, sub_3, sub_4, sub_5 = subject_marks
                    
                    try:
                        total_marks = driver.find_element(By.ID, "rptMarks_ctl06_lblTotal").text
                    except NoSuchElementException:
                        total_marks = driver.find_element(By.ID, "rptMarks_ctl07_lblTotal").text
                        
                    result = driver.find_element(By.ID, "lblresult").text

                    #Updating Excel Header and Subject
                    if self.format:
                        ws['A1'] = f"BCA Result Sem- {sem_header}"
                        col_headers = ["Sr.No", "Name", "Father Name", "Registration No", "Roll No", "", f"{sub_name1}", f"{sub_name2}", f"{sub_name3}", f"{sub_name4}", f"{sub_name5}", "", "Total Marks", "Result"]
                        for cols, header in enumerate(col_headers, start=1):
                            cols_char = get_column_letter(cols)
                            ws[f"{cols_char}3"] = header
                        self.format = False

                    data = [self.current_row - 3, student_name, father_name, reg_no, rows[1], "", sub_1, sub_2, sub_3, sub_4, sub_5, "", total_marks, result]
                    
                    for char in range(1, 15):
                        char_num = get_column_letter(char)
                        ws[f"{char_num}{self.current_row}"] = data[char - 1]
                    
                    # update the tracker file once per row
                    with open(self.row_tracker_file, "w") as file2:
                        file2.write(str(self.current_row + 1))
                        
                    self.current_row += 1

                    # Update GUI
                    self.progress_status += self.one_time_increment
                    self.progressbar.set(self.progress_status)
                    self.percentage_label.configure(text=f"{(self.progress_status)*100:.2f}%")
                    self.est_time_label.configure(text=f"Estimated Time: {max(0, self.estimate_time//4)} Minutes")
                    self.estimate_time -= 1

                    # Update processed entries and refresh
                    file.write(reg_no + "\n")
                    self.process_data.append(reg_no) 
                    driver.get("https://result.mdu.ac.in/postexam/result.aspx")
        # saving worked file when error occured while processing
        except Exception as e:
            wb.save(self.output_file.get())
            msg.showerror("Extraction Error", f"Error occurred during data extraction for reg: {rows[0]}.\nPlease start the process again with the same output file.\nDetails: {str(e)}")
            driver.quit()
            return

        # Save and Cleanup
        try:
            wb.save(self.output_file.get())        
        except PermissionError:
            msg.showerror("Permission Denied", f"Cannot save to {self.output_file.get()}.\n\nPlease ensure the Excel file is CLOSED and try again.")
            driver.quit()
            return

        with open(self.row_tracker_file, 'w') as track_update:
            track_update.write(str(self.current_row))
            
        driver.quit()
        msg.showinfo("Success", "All data fetched successfully.")
        
        if msg.askyesno("Clear Cache", "Process completed successfully. Clear cache files?"):
            cache_files = ['Processed_entries.txt', 'row_status.txt']
            for files in cache_files:
                if os.path.exists(files):
                    os.remove(files)
                    
        if self.output_file.get():
                try:
                    if platform.system() == "Windows":
                        os.startfile(self.output_file.get())
                    elif platform.system() == "Linux":
                        subprocess.run(["xdg-open",self.output_file.get()])
                    else:
                        subprocess.run(["open",self.output_file.get()])
                except Exception as e:
                    msg.showerror("Error",f"Unable to open file \n {str(e)}")
        else:
            if msg.askyesno("Open File","Would you like to open file ?"):
                try:
                    if platform.system() == "Windows":
                        os.startfile(self.output_file.get())
                    elif platform.system() == "Linux":
                        subprocess.run(["xdg-open",self.output_file.get()])
                    else:
                        subprocess.run(["open",self.output_file.get()])
                except Exception as e:
                    msg.showerror("Error",f"Unable to open file \n {str(e)}")

if __name__ == "__main__":
    app = AppResult()